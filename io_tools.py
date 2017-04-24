"""
Module for Input Output
"""

import openpyxl as pxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from typing import Union, Iterable
from cytoolz import isiterable


def read_sheet(sheet: Worksheet, *, header=True, index_col: Union[int, Iterable]=None) -> pd.DataFrame:
    """
    Read a openpyxl Worksheet object and return the sheet's data

    >>> wb = pxl.load_workbook("test.xlsx")
    >>> df = read_sheet(wb.active, index_col=0)

    :param Worksheet sheet: the sheet object
    :param bool header: if the sheet's first row represent the header, if header is True then colnames are converted to upper case strings
    :param Union[int, Iterable] index_col: the cols used as index, default no columns used as index
    """
    data: Iterable = sheet.values
    col_names = [str(col_name).upper() for col_name in next(data)] if header else None
    data = list(data)
    df = pd.DataFrame(data, columns=col_names)
    if isiterable(index_col):
        index_header = [col_names[i] for i in index_col] if header else index_col
        df.set_index(index_header, inplace=True)
    elif index_col is not None:
        index_header = col_names[index_col] if header else index_col
        df.set_index(index_header, inplace=True)
    else:
        pass
    return df


if __name__ == "__main__":
    from actuarial_tools import *
    wb = pxl.load_workbook("test.xlsx", data_only=True)
    ws = wb.active
    df = read_sheet(ws, index_col=0)
    print(df)
    print(df.values)

    wb = pxl.load_workbook("data/ProbabilityTables.xlsx", data_only=True)
    ws = wb.get_sheet_by_name("CL13_2")
    df = read_sheet(ws, index_col=0)
    # print(df.iloc[1:23,0])
    # print(df.values)
    pt = ProbabilityTable.from_dataframe(df, cat=DeathBenefit)
    mp = ModelPoint(sex=0, age=10, policy_term=10, payment_term=5, policy_year=2)
    print(pt(mp, time_scale=YEAR))
