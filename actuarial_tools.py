"""
This module defined tools for actuarial computation, including:

..py:class:: ModelPoint 模型点
..py:class:: TimeScale 现金流的时间类型， 年度还是月度
..py:class:: CashFlowType 现金计算类型, 与何相关
..py:class:: CashFlowSA 与保额相关现金流
..py:class:: CashFlowPrem 与保费相关现金流

...

"""
import calendar
from abc import ABCMeta, abstractmethod
from datetime import date
from enum import Flag, auto, Enum
from typing import *
from copy import deepcopy
from collections import OrderedDict
from functools import reduce
import xlwings as xw
from numpy import ndarray, concatenate, zeros, ones, array, full, repeat
import pandas as pd


class ModelPoint:
    """
        ModelPoint Class, represent a model point

        非空成员:
            - sex
            - age
            - policy_term
            - payment_term

        可选成员:
            - policy_year
            - policy_month

        用于index的计算成员:
            - index_policy_year
            - index_policy_month

        其它计算成员:
            - age_now

        特殊成员:
            - gross_premium
            - sum_assured

        重载操作符:
            - []

        特殊成员储存在 受保护的成员 *_data_pack* 中，*_data_pack* 本身是一个dict，模型点可以通过[]来访问它的 *_data_pack*、
        添加新的特殊成员
        
        >>> mp = ModelPoint(0, 10, 30, 5)
        >>> mp['gross_premium'] = 1000.0
        >>> mp.gross_premium
        1000.0
        >>> mp['cv'] = 0
        >>> mp.cv 
        0
        >>> del mp.cv

    """

    def __init__(self, sex: int, age: int, policy_term: Union[str, int], payment_term: Union[str, int], policy_year: Optional[int]=None, policy_month: Optional[int]=None, **kwargs):  # **kwargs 表示可变长度参数 key word arguments
        """
        建立一个新的模型点

        Example:

        >>> mp1 = ModelPoint(0, 10, 30, 5)
        >>> mp2 = ModelPoint(0, 10, '30@', '15')
        >>> mp3 = ModelPoint(0, 10, '@30', '15@', gross_premium=10.0)


        >>> ModelPoint(0, 10, 30, 5).sex
        0
        >>> ModelPoint(0, 10, 30, 5).payment_term
        5
        >>> ModelPoint(0, 10, '@30', '5@').policy_term
        20


        :param int sex: 性别
        :param int age: 年龄
        :param Union[str, int] policy_term: 保险期间 若为str 则须为 '20'， '@20'， '20@' 之一
        :param Union[str, int] payment_term: 缴费期间 格式同 policy_term
        :param Optional[int] policy_year: 保单年度 1 <= policy_year
        :param Optional[int] policy_month: 保单月度 1 <= policy_month < 12
        """

        # 变量检查：
        assert policy_year is None or 1 <= policy_year
        assert policy_month is None or 1 <= policy_month <= 12

        # 生成 对象属性（instance attribute)，即定义在对象中的变量
        self.sex: int = sex
        """ 性别 """  # 紧接着一个 对象成员变量 的字符串为 这个 对象成员变量 的 文档 （docstring）， 可以试着 ctrl + Q 看看
        self.age: int = age
        """ 年龄 """
        self._raw_policy_term = policy_term  # 以下划线开头的变量 表示 受保护的变量（即不建议被从外部访问）
        """ 未处理的保险期间 """
        self._raw_payment_term = payment_term
        """ 未处理的缴费期间 """
        self.policy_term: int = self.convert_term(policy_term, age)
        """ 保险期间 """
        self.payment_term: int = self.convert_term(payment_term, age)
        """ 缴费期间 """
        self.policy_year: int = policy_year
        """ 保单年度 """
        self.policy_month: int = policy_month
        """ 保单月度  """
        self._data_pack: Dict[str, Any] = kwargs
        """ 记录计算结果的地方 """

    def __getitem__(self, item):
        return self._data_pack[item]

    def __setitem__(self, key, value):
        self._data_pack[key] = value

    def __delitem__(self, key):
        if key in self._data_pack:
            del self._data_pack[key]

    def __getattribute__(self, name):
        try:
            return super().__getattribute__(name)
        except AttributeError as e:
            return self[name]

    def __delattr__(self, item):
        try:
            super().__delattr__(item)
        except AttributeError:
            del self[item]

    @staticmethod
    def convert_term(term: Union[str, int], age: Optional[int]=None) -> int:
        """
        将 字符类型 的 term 转化为 数字， 如 10 岁 保至 30 岁 则返回 20, 当 term 非 至XX岁这一情况， age可以不提供

        Example:

        >>> ModelPoint.convert_term('20')
        20
        >>> ModelPoint.convert_term(20)
        20
        >>> ModelPoint.convert_term('30@', 10)
        20
        >>> ModelPoint.convert_term('@30', age=15)
        15

        :param Union[str, int] term: 某种期间
        :param Optional[int] age: 年龄， 为应对 至XX岁这一情况
        :return: 实际期间
        :rtype: int
        """

        if isinstance(term, str):   # isinstance(a, A) 可以得到 a 是否 是 A 的一个实例（instance）
            try:    # 关于 try 语句 请见异常处理
                return int(term.strip("@")) - age if '@' in term else int(term)
            except ValueError:
                raise ValueError(f"term must be an int or str represent a number like '10' , '10@', '@10', but now {term}")
            except TypeError:
                raise ValueError(f"when you have a '@' in term, age can't be None")
        else:
            return int(term)

    @property
    def index_policy_year(self) -> Optional[int]:
        """
        计算用于 slicing 的policy year

        Example:

        >>> mp = ModelPoint(0, 10, 5, 5, 2, 3)
        >>> mp.index_policy_year
        1

        >>> benefit = [1.0] * mp.policy_term
        >>> benefit_left = benefit[mp.index_policy_year:]
        >>> benefit_left
        [1.0, 1.0, 1.0, 1.0]

        :return: policy_year - 1
        """
        try:
            return self.policy_year - 1
        except TypeError:
            return None

    @property
    def index_policy_month(self) -> Optional[int]:
        """
        计算用于 slicing 的policy month

        Example:

        >>> mp = ModelPoint(0, 10, 5, 5, 2, 3)
        >>> mp.index_policy_month
        14

        >>> benefit_yr = [1.0] * mp.policy_term
        >>> from functools import reduce
        >>> benefit_mth = reduce(lambda a, b: a + b, ([b]*12 for b in benefit_yr))
        >>> left_benefit_of_month = benefit_mth[mp.index_policy_month:]

        :return: index_policy_year * 12 + policy_month - 1
        """
        try:
            return self.index_policy_year * 12 + self.policy_month - 1
        except TypeError:
            return None

    @property
    def age_now(self) -> int:
        """ 当前保单年度的被保险人年龄 """
        return self.age + self.index_policy_year

    @staticmethod
    def month_delta(from_date: date, to_date: date) -> int:
        """
        计算两个日期之间的月份间隔

        :param date from_date:
        :param date to_date:
        :return: month num from from_date to to_date
        :rtype: int
        """
        days_in_month = calendar.monthrange(to_date.year, to_date.month)[1]
        imaginary_day_2 = 31 if to_date.day == days_in_month else to_date.day
        month_delta = (to_date.month - from_date.month) + (to_date.year - from_date.year) * 12 + \
                      (-1 if from_date.day > imaginary_day_2 else 0) + 1
        return month_delta

    def copy(self):
        """
        返回自身的一个深拷贝
        :rtype: ModelPoint
        """
        return deepcopy(self)

    def copy_as_initiate(self):
        """
        返回自身的一个深拷贝 只是保单时刻为发单时刻

        :rtype: ModelPoint
        """
        mp = self.copy()
        mp.policy_month = 1
        mp.policy_year = 1
        return mp

    @property
    def gross_premium(self) -> Optional[float]:
        """
        毛保费

        Example:

        >>> ModelPoint(0, 10, 30, 5, gross_premium=1.25).gross_premium
        1.25
        >>> ModelPoint(0, 10, 30, 5).gross_premium

        """
        try:
            return self._data_pack['gross_premium']
        except KeyError:
            return None

    @gross_premium.setter
    def gross_premium(self, val: float):
        """ setter of gross premium """
        assert val > 0
        self._data_pack['gross_premium'] = val

    @gross_premium.deleter
    def gross_premium(self):
        del self._data_pack['gross_premium']

    @property
    def premium_paid(self) -> ndarray:
        """
        已缴保费， 若模型点已有保费数据则返回当前保单保单年度以后的已交保费数据。
        若模型点未提供保单年度则默认为发单时刻
        """
        try:
            rst: ndarray = self.gross_premium * concatenate((ones(self.payment_term), zeros(self.policy_term - self.payment_term))).cumsum()[self.index_policy_year:]
            return rst
        except TypeError:
            return concatenate((ones(self.payment_term), zeros(self.policy_term - self.payment_term))).cumsum()[self.index_policy_year:]

    @property
    def sum_assured(self) -> Optional[float]:
        """ 保额 """
        try:
            return self._data_pack['sum_assured']
        except KeyError:
            return None

    @sum_assured.setter
    def sum_assured(self, val):
        self._data_pack['sum_assured'] = val

    @sum_assured.deleter
    def sum_assured(self):
        del self._data_pack['sum_assured']


class TimeScale(Enum):
    YEAR = auto()
    MONTH = auto()

    @classmethod
    def from_str(cls, str):
        """
        get obj from string 由于excel难以传入python obj 依靠 此函数 完成 python 与 excel 的桥接

        :param str:
        :return:
        """
        return getattr(cls, str)


YEAR = TimeScale.YEAR
MONTH = TimeScale.MONTH


class CashFlowType(metaclass=ABCMeta):
    """
     现金流计算方式类型.
     在绝大部分情形下，现金流是某个保单属性的比例，如已交保费、保额等，每一种计算方式类型对应了一个CashFlowType的子类，
     在建立一个具体的现金流时，需通过选择合适的 CashFlowType 的子类来 确定 该现金流是哪个保单属性的比例。通过提供 ratio 来 给出 一个具体的一致的比例值。
     CashFlowType 的 getCashFlow 通过 ratio 计算出 具体的 现金流

     >>> cf = DeathBenefit(CashFlowSA(ratio=0.1))

    """

    DEFAULT_TIME = None

    def __init_subclass__(cls, **kwargs):
        """
        收集子类信息

        >>> class A(CashFlowType):
        ...     id=10
        >>> A.__name__
        'A'
        >>> CashFlowType.A.id
        10

        :param kwargs: 冗余参数
        """
        super().__init_subclass__(**kwargs)
        if hasattr(CashFlowType, cls.__name__):
            raise ValueError("This Cash Flow Type Already Exists")
        else:
            setattr(CashFlowType, cls.__name__, cls)

    def __get__(self, instance, owner):
        """ 防止实例及子类取得其他子类 """
        if owner is CashFlowType and instance is None:
            return self
        else:
            raise TypeError("Only CashFlowType.SubLiabilityType is legal")

    def __init__(self, ratio: float=None):
        """
        :param ratio:  责任占比
        """
        self.ratio = ratio
        """ 基础的比例 """

    def getCashFlow(self, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH) -> ndarray:
        """
        通过 self.ratio 返回模型点在此责任下的 **不考虑概率、未贴现** 的现金流

        :param ModelPoint mp: 模型点
        :param bool forMonth: 是否返回月度现金流 默认 True
        :param bool from_init: 是否从发单时刻算起 默认 False
        :rtype: ndarray
        """
        return self.getCashFlowBase(mp=mp, from_init=from_init, time_scale=time_scale) * self.ratio

    @abstractmethod
    def getCashFlowBase(self, *, mp: ModelPoint, from_init: bool=False, time_scale: TimeScale=MONTH) -> ndarray:
        """
        cashflow = ratio * cash_flow_base 此函数计算 cash_flow_base 根据不同的 CashFlowType 请重载此函数

        :param ModelPoint mp: 模型点
        :param bool forMonth: 是否返回月度现金流 默认 True
        :param bool from_init: 是否从发单时刻算起 默认 False
        :rtype: ndarray
        """
        pass


class CashFlowSA(CashFlowType):
    """ 与保额成比例的现金流类型 """

    def getCashFlowBase(self, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH) -> ndarray:
        """
        :param ModelPoint mp: 模型点
        :param TimeScale time_scale: 返回月度现金流还是年度现金流 默认 MONTH
        :param bool from_init: 是否从发单时刻算起 默认 False
        :rtype: ndarray
        """
        forMonth = time_scale is MONTH
        try:
            yearRatio = full(mp.policy_term, mp.sum_assured)
        except TypeError:
            yearRatio = ones(mp.policy_term)
        if from_init:
            return repeat(yearRatio, 12) if forMonth else yearRatio
        else:
            return repeat(yearRatio, 12)[mp.index_policy_month:] if forMonth else yearRatio[mp.index_policy_year:]


class CashFlowPrem(CashFlowType):
    """ 与 **已交保费** 成比例的现金流类型 """

    def getCashFlowBase(self, mp: ModelPoint, *, from_init: bool = False, time_scale: TimeScale=MONTH) -> ndarray:
        """
        :param ModelPoint mp: 模型点
        :param TimeScale time_scale: 返回月度现金流还是年度现金流 默认 MONTH
        :param bool from_init: 是否从发单时刻算起 默认 False
        :rtype: ndarray
        """
        forMonth = time_scale is MONTH
        try:
            paid_prem = mp.gross_premium * concatenate(
                (ones(mp.payment_term), zeros(mp.policy_term - mp.payment_term))).cumsum()
        except TypeError:
            paid_prem = concatenate(
                (ones(mp.payment_term), zeros(mp.policy_term - mp.payment_term))).cumsum()
        if from_init:
            return repeat(paid_prem, 12) if forMonth else paid_prem
        else:
            return repeat(paid_prem, 12)[mp.index_policy_month:] if forMonth else paid_prem[mp.index_policy_year:]


class CashFlowGetterFactory:
    """
    用来生产 满足条件的 getCashFlow 函数 的帮助类
    """
    PAYMENT_TREM = "PaymentTerm"
    """  """
    AGE = "AGE"
    PAYMENT_TREM_AGE = "PaymentTerm_Age"
    DICT_TYPE = (PAYMENT_TREM, AGE, PAYMENT_TREM_AGE)
    """ 支持的字典存储方式的类型 """

    def __init__(self, dict_type: str, *, ratio_dict: dict=None, ratio_data_frame: pd.DataFrame=None):
        """
        >>> import openpyxl as pxl
        >>> from io_tools import read_sheet
        >>> wb = pxl.load_workbook("test.xlsx", data_only=True)  # data_only=True force excel formula as there computed values
        >>> ws = wb.active
        >>> df = read_sheet(ws, index_col=[0,1])
        >>> getCashFlow = CashFlowGetterFactory(CashFlowGetterFactory.PAYMENT_TREM_AGE, ratio_data_frame=df)
        """
        assert dict_type in self.DICT_TYPE
        assert ratio_dict is not None or ratio_data_frame is not None
        if ratio_dict is not None:
            self.ratioDict = self._forceKeyAsInt_(ratio_dict)
        else:
            self.ratioDict = self._forceKeyAsInt_(self._genRatioDictFromDataFrame_(df=ratio_data_frame))
        """ 比例字典 """
        self.dictType = dict_type
        """ 比例字典存储的方式 """

    @classmethod
    def _genRatioDictFromDataFrame_(cls, df: pd.DataFrame) -> dict:
        index = df.index
        try:
            generator = ((k, cls._genRatioDictFromDataFrame_(df.loc[k])) for k in index.levels[0])
        except (AttributeError, IndexError):
            generator = ((k, df.loc[k].values) for k in index.values)
        return dict(generator)

    @classmethod
    def _forceKeyAsInt_(cls, d: dict) -> dict:
        """
        将字典类型的key转换为int

        :param dict d: 存储比例的字典
        :return: 键为整数的存储比例的字典
        """
        if not isinstance(d, dict):
            return d
        else:
            couple = [(int(k), cls._forceKeyAsInt_(v)) for k, v in d.items()]
            return dict(couple)

    def __call__(self, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH):
        """
        根据模型点计算现金流比例

        :param ModelPoint mp: 模型点
        :param bool from_init: 是否从发单时刻开始
        :param TimeScale time_scale: 结果的时间尺度
        :return: 现金流比例
        """
        if self.dictType == self.AGE:
            raw = self.ratioDict[mp.age]
        elif self.dictType == self.PAYMENT_TREM:
            raw = self.ratioDict[mp.payment_term]
        else:
            raw = self.ratioDict[mp.payment_term][mp.age]
        length = len(raw)
        yearRatio = array(raw)[:mp.policy_term] if mp.policy_term <= length else concatenate((raw, full(mp.policy_term - length, raw[-1])))
        if time_scale is YEAR:
            return yearRatio if from_init else yearRatio[mp.index_policy_year:]
        elif time_scale is MONTH:
            return repeat(yearRatio, 12)[mp.index_policy_month:] if not from_init else repeat(yearRatio, 12)


class __CashFlowCatMeta__(ABCMeta):

    def __contains__(cls, item):
        """
        判断一个现金流类型是否属于另一个现金流类型

        :param item: a class object or a instance
        :return:
        """
        try:
            return issubclass(item, cls) or isinstance(item, cls)
        except TypeError:
            return isinstance(item, cls)


class CashFlowCat(metaclass=__CashFlowCatMeta__):
    """ 现金流类型 """

    DEFAULT_TIME = None
    """ 默认现金流发生时间 """

    def __init_subclass__(cls, **kwargs):
        super().__init_subclass__(**kwargs)
        if hasattr(CashFlowCat, cls.__name__):
            raise ValueError("This Cash Flow Cat Already Exists")
        else:
            setattr(CashFlowCat, cls.__name__, cls)

    @classmethod
    def __contains__(cls, item):
        """
        判断一个现金流类型是否属于另一个现金流类型

        >>> issubclass(SurvivalBenefit, Benefit)
        True
        >>> SurvivalBenefit in Benefit
        True
        >>> SurvivalBenefit in Benefit(CashFlowSA(0.1))
        True
        >>> SurvivalBenefit(CashFlowSA(0.1)) not in Benefit(CashFlowSA(0.1))
        False

        :param item: a class object or a instance
        """
        try:
            return issubclass(item, cls) or isinstance(item, cls)
        except TypeError:
            return isinstance(item, cls)

    def __get__(self, instance, owner):
        """ 防止实例取得其他子类 """
        if issubclass(owner, CashFlowCat) and instance is None:
            return self
        elif isinstance(owner, ProductManager):
            return self
        else:
            raise TypeError("Only CashFlowType.SubLiabilityType is legal")

    def __init__(self, cash_flow_type: CashFlowType=None, *, time=None, getCashFlow: Callable=None, index_in_product: int=None):
        """
        :param CashFlowType cash_flow_type:
        :param float time: 现金发生时点, 范围 为 区间 [0,1]
        :param Callable getCashFlow: 自定义的getCashFlow函数，见CashFlowGetterFactory类
        :param int index_in_product: 在产品中的编号, 默认由 ProductManager 根据定义顺序 初始化
        """
        # assert cash_flow_type.ratio is not None or getCashFlow is not None
        self.type = cash_flow_type
        """ 现金流基础类型 """
        self.time = time if time is not None else self.DEFAULT_TIME
        """ 现金流产生的时刻 """
        self.index_in_product: Optional[int] = index_in_product
        """ 在产品中的编号"""
        self.cashFlowGetter = getCashFlow

    def getCashFlow(self, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH) -> ndarray:
        try:
            return self.cashFlowGetter(mp=mp, from_init=from_init, time_scale=time_scale) * self.type.getCashFlowBase(mp=mp, from_init=from_init, time_scale=time_scale)
        except TypeError:
            return self.type.getCashFlow(mp=mp, from_init=from_init, time_scale=time_scale)

    @classmethod
    def from_str(cls, str):
        return getattr(cls, str)


# 定义CashFlow 为 CashFlowCat 别名
CashFlow = CashFlowCat


class Benefit(CashFlow):
    """
    保险责任现金流
    """
    DEFAULT_TIME = 0.5


class DeathBenefit(Benefit):
    """
    死亡保险责任现金流
    """
    pass


class SurvivalBenefit(Benefit):
    """
    生存保险责任
    """
    DEFAULT_TIME = 1.0


class MaturityBenefit(SurvivalBenefit):

    def getCashFlow(self, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH):
        rst = super().getCashFlow(mp=mp, from_init=from_init, time_scale=time_scale)
        rst[:-1] = 0
        return rst

class AccidentBenefit(Benefit):
    """
    意外保险责任
    """
    pass


class IllnessBenefit(Benefit):
    """
    疾病保险责任
    """
    pass


class CriticalIllnessBenefit(IllnessBenefit):
    """
    重大疾病保险责任
    """
    pass


class OtherIllnessBenefit(IllnessBenefit):
    """
    轻症保险责任
    """
    pass


class LapseCashFlow(CashFlow):
    """
    退保现金流
    """
    DEFAULT_TIME = 0.5

    def getCashFlow(self, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH) -> ndarray:
        # TODO: return cv array of the model point
        pass


class ProbabilityTable:
    """
    概率表
    """
    AGE = "AGE"
    POL_YR = "POL_YR"
    TABLES = []

    def __init__(self, array: ndarray, cat: CashFlowCat, type=None, table_name=None):
        # assert 2 in array.shape
        assert type is None or type in (self.AGE, self.POL_YR)
        self.values = array if array.shape[0] < array.shape[1] else array.T
        self.cat = cat
        self.type = self.AGE if type is None else type
        self.table_name = table_name
        self.TABLES.append(self)
        self.table_index = len(self.TABLES)

    @classmethod
    def from_dataframe(cls, df: pd.DataFrame, *, cat: CashFlowCat, type=AGE, table_name=None):
        """
        Create a ProbabilityTable from dataframe

        >>> wb = pxl.load_workbook("data/ProbabilityTables.xlsx", data_only=True)
        >>> ws = wb.get_sheet_by_name("CL13_2")
        >>> df = read_sheet(ws, index_col=0)
        >>> pt = ProbabilityTable.from_dataframe(df,cat=DeathBenefit)
        """

        array = df.values
        type = df.index.name.upper() if type is None else type
        return ProbabilityTable(array=array, cat=cat, type=type, table_name=table_name)

    def __getitem__(self, item):
        return self.values[item]

    def __call__(self, mp: ModelPoint, from_init: bool=False, time_scale: TimeScale=MONTH, tuner: Callable[[ndarray], ndarray]=None):

        if self.type == self.AGE:
            probs = self[mp.sex, mp.age: mp.age + mp.policy_term]
        elif self.type == self.POL_YR:
            probs = self[mp.sex, mp.index_policy_year: mp.index_policy_year + mp.policy_term]
        else:
            raise TypeError
        # use the tuner to tune the probability
        try:
            probs = tuner(probs)
        except TypeError:
            pass

        if time_scale is MONTH:
            probs = repeat(probs, 12)
        elif time_scale is YEAR:
            pass
        else:
            raise ValueError(f"{time_scale}")

        if from_init:
            return probs[mp.index_policy_year:] if time_scale is YEAR else probs[mp.index_policy_month:]
        else:
            return probs


class ProbabilityPac:

    def __init__(self, *args, prob_pac_index):
        self.probabilityTables: List[ProbabilityTable] = args
        self.prob_pac_index = prob_pac_index

    def getProbability(self, cat):
        return filter(lambda x: x.cat in cat, self.probabilityTables)

    @property
    def death_probability_table(self) -> ProbabilityTable:
        return next(self.getProbability(DeathBenefit))

    @property
    def critical_illness_probability_table(self) -> ProbabilityTable:
        try:
            return next(self.getProbability(IllnessBenefit))
        except StopIteration:
            return None

    @property
    def accident_probability_table(self) -> Iterable[ProbabilityTable]:
        return self.getProbability(AccidentBenefit)

    def __call__(self, mp: ModelPoint, from_init: bool=False, time_scale=MONTH, tunners: Dict[CashFlow, Callable]=None) -> OrderedDict:
        # TODO: How to integrate TUNER ?
        pass

    @staticmethod
    def inForce(death_rate: ndarray, *, ci_rate: ndarray=None, ci_k: ndarray=None, lapse: ndarray=None, killAllAtEnd: bool=True)->ndarray:
        """
        计算Inforce at end of each period

        :param death_rate: 调整后的死亡率
        :param ci_rate: 调整后的疾病发生率
        :param ci_k: 调整后的疾病致死占比
        :param lapse: 调整后的退保率
        """
        if ci_k is not None:
            death_rate = death_rate * (1 - ci_k)
        if killAllAtEnd and ci_rate is not None:
            death_rate[-1] = 1 - ci_rate[-1]
        survival_rate = 1 - death_rate
        if ci_rate is not None:
            survival_rate -= ci_rate
        if lapse is not None:
            survival_rate *= (1 - lapse)
        if_eop = survival_rate.cumprod()
        return if_eop


class ProductManager(type):

    PRODUCTS = {}

    def __new__(mcs, name, bases, namespace, *args, prod_id=None, prod_name=None, **kwargs):
        if name in mcs.PRODUCTS:
            raise ValueError(f"Product {name} already exists")
        if prod_id is not None:
            namespace['prod_id'] = prod_id
        if prod_name is not None:
            namespace['prod_name'] = prod_name

        # init CashFlow's attribute ``index_in_product``
        cfs: List[CashFlow] = [x for x in namespace.values() if isinstance(x, CashFlow)]
        existed_cf_indexes = [x.index_in_product for x in cfs if x.index_in_product is not None]
        assert len(existed_cf_indexes) == len(set(existed_cf_indexes))
        for cf, idx in zip(filter(lambda x: x.index_in_product is None, cfs), [x for x in range(len(cfs)) if x not in existed_cf_indexes][:len(cfs) - len(existed_cf_indexes)]):
            cf.index_in_product = idx

        namespace["CashFlows"] = cfs
        # init cls
        cls = type.__new__(mcs, name, bases, namespace)
        if name != "ProductBase":
            mcs.PRODUCTS[cls.prod_id] = cls
        setattr(mcs, name, cls)
        return cls

    def __str__(cls):
        return f"{cls.prod_name} {cls.prod_id}"


class ProductBase(metaclass=ProductManager):

    CashFlows: List[CashFlow]
    """ 类成员， 保存了该产品的所有现金流 """

    def __get__(self, instance, owner):
        if instance is None and owner is ProductManager:
            return self

    # def __init__(self, probabilityTables: Iterable[ProbabilityTable]):
    #     self.probabilityTables = list(probabilityTables)

    @classmethod
    def getCashFlow(cls, cash_flow_cat: CashFlowCat, mp: ModelPoint, *, from_init: bool=False, time_scale: TimeScale=MONTH):
        return reduce(lambda x, y: x + y, map(lambda c: c.getCashFlow(mp, from_init=from_init, time_scale=time_scale), filter(lambda c: c in cash_flow_cat, cls.CashFlows)))


if __name__ == '__main__':
    import doctest
    import openpyxl as pxl
    from io_tools import read_sheet
    print(ProductManager.PRODUCTS)
    wb = pxl.load_workbook("test.xlsx", data_only=True)
    ws = wb.active
    df = read_sheet(ws, index_col=[0, 1])
    d = CashFlowGetterFactory._genRatioDictFromDataFrame_(df)
    print(d)
    doctest.testmod()
